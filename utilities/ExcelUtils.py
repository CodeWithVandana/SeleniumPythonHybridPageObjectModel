import openpyxl

def get_row_count(path,sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row

def get_column_count(path,sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_column

def get_cell_data(path,sheet_name,row_number,column_number):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_number,column=column_number).value

def set_cell_data(path,sheet_name,row_number,column_number,data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_number,column=column_number).value=data
    workbook.save(path)

def get_data_from_excel(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        email, password = row
        if email and password:   # skip empty rows
            data.append((str(email).strip(), str(password).strip()))
    return data
